"""Test e2e del reporte Excel: fixtures → parse → clasifica → calcula → xlsx."""
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

import calculo_impuestos
import clasificador
import parse_plataforma
from generar_reporte import genera_reporte

FIXTURES = Path(__file__).resolve().parent / "fixtures"


@pytest.fixture(scope="module")
def jsons(tmp_path_factory):
    carpeta = tmp_path_factory.mktemp("reporte")
    clas, plat, calc = carpeta / "clas.json", carpeta / "plat.json", carpeta / "calc.json"
    clasificador.main([str(FIXTURES / "cfdis.zip"), "--periodo", "2026-05", "-o", str(clas)])
    parse_plataforma.main([str(FIXTURES / "airbnb_ganancias_2026-05.csv"), "-o", str(plat)])
    calculo_impuestos.main([
        "--clasificacion", str(clas), "--plataforma", str(plat),
        "--periodo", "2026-05", "-o", str(calc),
    ])
    return carpeta, json.loads(clas.read_text(encoding="utf-8")), json.loads(calc.read_text(encoding="utf-8"))


def test_genera_xlsx_completo(jsons):
    carpeta, clasificacion, calculo = jsons
    ruta = genera_reporte(clasificacion, calculo, carpeta / "reporte.xlsx")
    wb = load_workbook(ruta)
    assert wb.sheetnames == [
        "Resumen", "Deducibles", "Dudosas", "Rechazadas",
        "Deducciones personales", "Inversiones",
    ]

    resumen = wb["Resumen"]
    textos = [str(c.value) for fila in resumen.iter_rows() for c in fila if c.value]
    assert any("FORMULARIO ISR" in t for t in textos)
    assert any("herramienta de apoyo" in t for t in textos)  # disclaimer presente

    deducibles = wb["Deducibles"]
    # 2 deducibles + 1 ajuste (nota de crédito) + encabezado
    assert deducibles.max_row == 4
    reglas = {deducibles.cell(row=r, column=12).value for r in range(2, 5)}
    assert reglas == {"H-01", "H-02", "R-06"} or reglas == {"H-01", "R-06"}

    personales = wb["Deducciones personales"]
    assert personales.max_row == 2  # la colegiatura D10
    assert personales.cell(row=2, column=6).value == "D10"
