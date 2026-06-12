"""parse_plataforma.py — Lee el reporte de ganancias de una plataforma y emite JSON normalizado.

Fase 1: adaptador Airbnb (CSV o XLSX). Diseño por adaptadores: para agregar
Uber/Didi/Mercado Libre se implementa otra clase con `detecta()` y `parsea()`
y se registra en ADAPTADORES — sin tocar el resto del flujo.

El adaptador Airbnb hereda la lógica validada del flujo legacy del proyecto:
- Detección tolerante de columnas por nombre normalizado (sin acentos).
- Filas de retención identificadas por la columna "Tipo":
  "Retención del impuesto sobre la renta para México" → ISR retenido
  "Retención del IVA en México" → IVA retenido
- parse_money tolerante a formatos $1,234.56 / 1.234,56 / (123.45).

Salida normalizada (montos Decimal, serializados como string en JSON):
{
  "plataforma": "airbnb", "archivo": ...,
  "totales": {"ingresos_recibidos", "isr_retenido", "iva_retenido", ...},
  "por_mes": {"2026-05": {...}},
  "advertencias": [...]
}

Uso:
    python parse_plataforma.py ganancias_airbnb.csv
    python parse_plataforma.py ganancias.xlsx -o salida.json
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path


def normaliza_texto(valor: object) -> str:
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def parse_money(valor: object) -> Decimal | None:
    """Convierte montos con símbolos, comas y paréntesis a Decimal. None si no es monto."""
    if valor is None:
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    crudo = str(valor).strip()
    if not crudo:
        return None

    limpio = crudo.replace("(", "-").replace(")", "")
    limpio = re.sub(r"[^0-9,.\-]", "", limpio)
    if not limpio or limpio in {"-", ".", ","}:
        return None

    if "," in limpio and "." in limpio:
        if limpio.rfind(",") > limpio.rfind("."):
            limpio = limpio.replace(".", "").replace(",", ".")
        else:
            limpio = limpio.replace(",", "")
    elif "," in limpio:
        partes = limpio.split(",")
        if len(partes) == 2 and len(partes[1]) in (1, 2):
            limpio = limpio.replace(",", ".")
        else:
            limpio = limpio.replace(",", "")
    elif "." in limpio:
        partes = limpio.split(".")
        if len(partes) > 2:
            limpio = "".join(partes[:-1]) + "." + partes[-1]

    try:
        return Decimal(limpio)
    except InvalidOperation:
        return None


def busca_columna(columnas: list[str], candidatos: list[str]) -> str | None:
    """Columna cuyo nombre normalizado coincide (exacto primero, luego contiene)."""
    mapa = {col: normaliza_texto(col) for col in columnas}
    for cand in candidatos:
        cand_n = normaliza_texto(cand)
        for col, col_n in mapa.items():
            if cand_n == col_n:
                return col
    for cand in candidatos:
        cand_n = normaliza_texto(cand)
        for col, col_n in mapa.items():
            if cand_n in col_n:
                return col
    return None


def _parse_fecha(valor: object) -> tuple[str | None, bool]:
    """Devuelve ("AAAA-MM", ambigua). Airbnb exporta MM/DD/YYYY por default."""
    if isinstance(valor, datetime):
        return f"{valor.year:04d}-{valor.month:02d}", False
    texto = str(valor or "").strip()
    if not texto:
        return None, False
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", texto)
    if m:
        return f"{m.group(1)}-{m.group(2)}", False
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", texto)
    if m:
        primero, segundo, anio = int(m.group(1)), int(m.group(2)), m.group(3)
        if primero > 12:  # solo puede ser DD/MM/YYYY
            return f"{anio}-{segundo:02d}", False
        if segundo > 12:  # solo puede ser MM/DD/YYYY
            return f"{anio}-{primero:02d}", False
        # Ambiguo: se asume MM/DD/YYYY (formato de exportación de Airbnb).
        return f"{anio}-{primero:02d}", True
    return None, False


class AdaptadorAirbnb:
    nombre = "airbnb"

    COLS_TIPO = ["Tipo", "Type"]
    COLS_MONTO = ["Monto", "Amount"]
    COLS_INGRESO = ["Ingresos recibidos", "Paid out", "Payout", "Earnings"]
    COLS_FECHA = ["Fecha", "Date"]

    # Frases reales de la columna "Tipo" del reporte de Airbnb México (validadas
    # en el flujo legacy con datos reales del autor).
    MARCA_ISR = ["retencion del impuesto sobre la renta para mexico", "retencion isr"]
    MARCA_IVA = ["retencion del iva en mexico", "retencion iva"]

    def detecta(self, columnas: list[str]) -> bool:
        return (
            busca_columna(columnas, self.COLS_TIPO) is not None
            and busca_columna(columnas, self.COLS_MONTO) is not None
            and busca_columna(columnas, self.COLS_INGRESO) is not None
        )

    def parsea(self, filas: list[dict], archivo: str) -> dict:
        columnas = list(filas[0].keys()) if filas else []
        col_tipo = busca_columna(columnas, self.COLS_TIPO)
        col_monto = busca_columna(columnas, self.COLS_MONTO)
        col_ingreso = busca_columna(columnas, self.COLS_INGRESO)
        col_fecha = busca_columna(columnas, self.COLS_FECHA)

        faltantes = [
            nombre for nombre, col in
            [("Tipo", col_tipo), ("Monto", col_monto), ("Ingresos recibidos", col_ingreso)]
            if col is None
        ]
        if faltantes:
            raise ValueError(
                f"No se identificaron columnas {faltantes}. Columnas disponibles: {columnas}"
            )

        advertencias: list[str] = []
        fecha_ambigua = False
        vacio = {
            "ingresos_recibidos": Decimal("0"),
            "isr_retenido": Decimal("0"),
            "iva_retenido": Decimal("0"),
            "filas_ingreso": 0,
            "filas_retencion": 0,
        }
        totales = dict(vacio)
        por_mes: dict[str, dict] = {}

        for fila in filas:
            if all(str(v or "").strip() == "" for v in fila.values()):
                continue
            tipo = normaliza_texto(fila.get(col_tipo))
            monto = parse_money(fila.get(col_monto))
            ingreso = parse_money(fila.get(col_ingreso))

            mes, ambigua = (None, False)
            if col_fecha:
                mes, ambigua = _parse_fecha(fila.get(col_fecha))
            fecha_ambigua = fecha_ambigua or ambigua
            clave_mes = mes or "sin_fecha"
            cubeta = por_mes.setdefault(clave_mes, dict(vacio))

            if ingreso is not None and ingreso != 0:
                totales["ingresos_recibidos"] += ingreso
                cubeta["ingresos_recibidos"] += ingreso
                totales["filas_ingreso"] += 1
                cubeta["filas_ingreso"] += 1

            if monto is None:
                continue
            if any(marca in tipo for marca in self.MARCA_IVA):
                totales["iva_retenido"] += abs(monto)
                cubeta["iva_retenido"] += abs(monto)
                totales["filas_retencion"] += 1
                cubeta["filas_retencion"] += 1
            elif any(marca in tipo for marca in self.MARCA_ISR):
                totales["isr_retenido"] += abs(monto)
                cubeta["isr_retenido"] += abs(monto)
                totales["filas_retencion"] += 1
                cubeta["filas_retencion"] += 1

        if fecha_ambigua:
            advertencias.append(
                "Fechas DD/MM vs MM/DD ambiguas en el reporte; se asumió MM/DD/AAAA "
                "(formato de exportación de Airbnb). Verifica los meses del resumen."
            )
        if not col_fecha:
            advertencias.append("El reporte no trae columna de fecha; totales sin desglose mensual.")

        return {
            "plataforma": self.nombre,
            "archivo": archivo,
            "columnas_usadas": {
                "tipo": col_tipo, "monto": col_monto,
                "ingreso": col_ingreso, "fecha": col_fecha,
            },
            "totales": totales,
            "por_mes": dict(sorted(por_mes.items())),
            "advertencias": advertencias,
        }


ADAPTADORES = [AdaptadorAirbnb()]


def _lee_filas(ruta: Path) -> list[dict]:
    """CSV (utf-8 con o sin BOM) o XLSX (primera hoja) → lista de dicts por fila."""
    if ruta.suffix.lower() == ".csv":
        with open(ruta, encoding="utf-8-sig", newline="") as f:
            return [dict(fila) for fila in csv.DictReader(f)]
    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Para leer Excel se requiere openpyxl (pip install openpyxl); "
                "alternativa: exporta el reporte como CSV."
            ) from exc
        wb = load_workbook(ruta, read_only=True, data_only=True)
        hoja = wb.worksheets[0]
        filas_iter = hoja.iter_rows(values_only=True)
        encabezados = [str(c).strip() if c is not None else "" for c in next(filas_iter, [])]
        filas = []
        for valores in filas_iter:
            filas.append({encabezados[i]: valores[i] if i < len(valores) else None
                          for i in range(len(encabezados)) if encabezados[i]})
        wb.close()
        return filas
    raise ValueError(f"Formato no soportado: {ruta.suffix} (se espera .csv o .xlsx)")


def parse_reporte(ruta: str | Path, plataforma: str | None = None) -> dict:
    """Punto de entrada programático."""
    ruta = Path(ruta)
    filas = _lee_filas(ruta)
    if not filas:
        raise ValueError(f"El reporte está vacío: {ruta}")
    columnas = list(filas[0].keys())

    candidatos = ADAPTADORES
    if plataforma:
        candidatos = [a for a in ADAPTADORES if a.nombre == plataforma]
        if not candidatos:
            disponibles = [a.nombre for a in ADAPTADORES]
            raise ValueError(f"Plataforma '{plataforma}' no soportada. Disponibles: {disponibles}")

    for adaptador in candidatos:
        if adaptador.detecta(columnas):
            return adaptador.parsea(filas, ruta.name)

    raise ValueError(
        "Ningún adaptador reconoce las columnas del reporte. "
        f"Columnas encontradas: {columnas}. "
        "Si es un reporte de Airbnb, exporta el CSV de ganancias completo "
        "(debe incluir las columnas Tipo, Monto e Ingresos recibidos)."
    )


def _json_default(value):
    if isinstance(value, Decimal):
        return str(value)
    raise TypeError(f"No serializable: {type(value)}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Parser de reportes de plataformas (Airbnb) a JSON")
    parser.add_argument("reporte", help="CSV o XLSX exportado de la plataforma")
    parser.add_argument("-o", "--salida", help="Archivo JSON de salida (default: stdout)")
    parser.add_argument("--plataforma", help="Forzar adaptador (default: autodetección)")
    args = parser.parse_args(argv)

    resultado = parse_reporte(args.reporte, args.plataforma)
    texto = json.dumps(resultado, ensure_ascii=False, indent=2, default=_json_default)
    if args.salida:
        Path(args.salida).write_text(texto, encoding="utf-8")
    else:
        print(texto)

    t = resultado["totales"]
    print(
        f"[parse_plataforma] {resultado['plataforma']}: ingresos {t['ingresos_recibidos']}, "
        f"ISR retenido {t['isr_retenido']}, IVA retenido {t['iva_retenido']}",
        file=sys.stderr,
    )
    for adv in resultado["advertencias"]:
        print(f"[parse_plataforma]   aviso: {adv}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
