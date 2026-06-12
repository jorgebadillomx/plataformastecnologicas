"""generar_reporte.py — Genera el Excel de resumen mensual a partir de la clasificación y el cálculo.

Hojas:
  Resumen                qué capturar en el SAT + conciliación + disclaimer
  Deducibles             facturas aceptadas (y ajustes) con regla citada
  Dudosas                pendientes de decisión del usuario, con motivo
  Rechazadas             no deducibles y diferidas, con regla y motivo
  Deducciones personales candidatas para la declaración ANUAL (UsoCFDI D0x)
  Inversiones            activo fijo detectado (UsoCFDI I0x), se reporta sin depreciar

Uso:
    python generar_reporte.py --clasificacion clas.json --calculo calc.json -o reporte.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NEGRITA = Font(bold=True)
ENCABEZADO_FILL = PatternFill("solid", fgColor="1F4E78")
ENCABEZADO_FONT = Font(bold=True, color="FFFFFF")
MONEDA_FMT = "#,##0.00"

COLUMNAS_FACTURA = [
    ("Fecha", "fecha", 20),
    ("Emisor", "emisor", 38),
    ("RFC emisor", "emisor_rfc", 15),
    ("Descripción", "descripcion", 50),
    ("UUID", "uuid", 38),
    ("Uso CFDI", "uso_cfdi", 9),
    ("Forma pago", "forma_pago", 10),
    ("Total", "total", 12),
    ("IVA detectado", "iva_detectado", 13),
    ("IVA acreditable", "iva_acreditable", 14),
    ("Base deducible ISR", "base_deducible_isr", 17),
    ("Regla", "regla", 9),
    ("Motivo", "motivo", 60),
]

CAMPOS_MONTO = {"total", "iva_detectado", "iva_acreditable", "base_deducible_isr"}

# Consejos accionables por regla (ID → texto). Clave "" = fallback genérico.
_CONSEJOS: dict[str, str] = {
    # ── Heurísticas (DUDOSAS) ──────────────────────────────────────────────────
    "H-01": (
        "Comisión/servicio de plataforma (Airbnb, Uber…). Normalmente DEDUCIBLE. "
        "Incluye si es una comisión o cargo de la plataforma por tu actividad."
    ),
    "H-02": (
        "Servicio del inmueble rentado (luz, agua, gas, internet, limpieza, "
        "mantenimiento). Incluye si el servicio corresponde al departamento que "
        "rentas en Airbnb. Si también es tu residencia, solo aplica la parte "
        "proporcional (fase 2)."
    ),
    "H-05": (
        "Compra en marketplace (Amazon, Walmart, Costco, etc.). "
        "Incluye si el artículo es para el departamento Airbnb (blancos, utensilios, "
        "artículos de decoración, herramientas de mantenimiento). "
        "Excluye si es para uso personal."
    ),
    "H-06": (
        "Internet o telefonía: uso mixto personal/Airbnb. "
        "Decide: ¿lo usas principalmente para administrar tu Airbnb? → Incluye. "
        "¿Es principalmente personal? → Excluye. "
        "En fases futuras se calculará el prorrateo exacto."
    ),
    "H-07": (
        "Combustible o gasto de vehículo. Para hospedaje: deducible solo si el "
        "vehículo se usa directamente para la operación (p. ej. traslado de insumos "
        "al inmueble). Si es para uso personal → excluye."
    ),
    "H-08": (
        "Sin categoría clara. Revisa si el gasto tiene relación directa y exclusiva "
        "con tu actividad Airbnb. Si puedes justificarlo ante el SAT → incluye; "
        "si hay dudas → excluye para no tener problemas."
    ),
    # ── Reglas duras — versión DUDOSA (monto bajo / revisable) ────────────────
    "R-03": (
        "Pago en efectivo ≤ $2,000. Deducible SOLO si el gasto es estrictamente "
        "necesario para tu actividad Airbnb. Combustible en efectivo NUNCA es "
        "deducible (LISR 27-III). Para el futuro: paga con tarjeta o transferencia."
    ),
    "R-09": (
        "Factura en moneda extranjera. Para acreditarla necesitas el tipo de cambio "
        "oficial del día del CFDI (Diario Oficial de la Federación). Si lo tienes, "
        "puedes incluirla y convertirla a pesos."
    ),
    "R-21": (
        "Activo fijo (inversión). No se deduce como gasto directo en el mes — se "
        "deprecia anualmente (porcentajes LISR Arts. 33-35). Aparece en la hoja "
        "'Inversiones' para tu declaración anual. Consulta a tu contador el % de "
        "depreciación que aplica a este bien."
    ),
    "R-23": (
        "UsoCFDI 'Sin efectos fiscales' (S01). Si el gasto es de tu actividad Airbnb, "
        "pide al proveedor que cancele esta factura y la reexpida con UsoCFDI G03 "
        "(Gastos en general). Una vez reexpedida, vuelve a correr el skill."
    ),
    # ── Reglas duras — versión NO_DEDUCIBLE / DIFERIDA ────────────────────────
    "R-02": (
        "Pago en efectivo > $2,000. No tiene solución retroactiva (LISR Art. 27-III). "
        "Esta factura no es deducible ni genera IVA acreditable. "
        "Para el futuro: paga siempre con tarjeta de débito/crédito o transferencia bancaria."
    ),
    "R-04": (
        "Factura PPD sin complemento de pago en este mes (DIFERIDA). No es un rechazo: "
        "se acreditará en el mes en que la pagaste y el proveedor emitió el CFDI "
        "de pago (complemento). Vuelve a correr el skill con el ZIP del mes en que "
        "realizaste el pago."
    ),
    "R-07": (
        "El RFC receptor de esta factura no coincide con el tuyo. Esta factura "
        "no es tuya fiscalmente. Pide al proveedor que la cancele y reexpida a tu RFC."
    ),
    "R-08": (
        "Factura cancelada. No tiene validez fiscal (CFF 29-A). "
        "Si el gasto fue real, pide al proveedor que te emita una nueva factura vigente."
    ),
    "R-14": (
        "Régimen de pagos definitivos. En esta opción (ficha 4/PLT) no existen "
        "deducciones ni IVA acreditable — el SAT toma la retención como pago final. "
        "Si quieres deducir gastos, deberías estar en el régimen de pagos provisionales "
        "(consulta a tu contador antes de cambiar, es irrevocable 5 años)."
    ),
    "R-22": (
        "Deducción personal (UsoCFDI D0x: médico, colegiatura, funeraria, etc.). "
        "No es gasto de la actividad Airbnb, pero sí puede reducir tu ISR en la "
        "declaración ANUAL. Aparece en la hoja 'Deducciones personales'."
    ),
    "H-03": (
        "Gasto financiero (banco, intereses, seguros, comisiones bancarias). "
        "No es deducible como gasto de la actividad Airbnb (no es estrictamente "
        "indispensable para la operación de hospedaje)."
    ),
    "H-04": (
        "Gasto personal (salud, ropa, supermercado, farmacia). No es gasto de "
        "la actividad Airbnb. Si es médico, colegiatura u otro concepto D0x, "
        "puede ser deducción personal en tu declaración ANUAL."
    ),
    "": (
        "Revisa el motivo de la columna anterior. Si crees que aplica a tu "
        "actividad Airbnb y puedes justificarlo ante el SAT, usa decisiones.json "
        "para incluirla en el siguiente ciclo del clasificador."
    ),
}


def _consejo(regla: str, estatus: str) -> str:
    texto = _CONSEJOS.get(regla, "")
    if not texto:
        texto = _CONSEJOS[""]
    return texto


def _num(valor) -> float | str:
    try:
        return float(Decimal(str(valor)))
    except Exception:
        return str(valor)


def _hoja_facturas(wb: Workbook, titulo: str, evaluaciones: list[dict],
                   add_consejo: bool = False) -> None:
    ws = wb.create_sheet(titulo)
    columnas = list(COLUMNAS_FACTURA)
    if add_consejo:
        columnas = columnas + [("¿Qué puedo hacer?", "_consejo", 72)]
    for col, (nombre, _, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col, value=nombre)
        celda.font = ENCABEZADO_FONT
        celda.fill = ENCABEZADO_FILL
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for fila, ev in enumerate(evaluaciones, start=2):
        for col, (_, clave, _) in enumerate(columnas, start=1):
            if clave == "_consejo":
                valor = _consejo(ev.get("regla", ""), ev.get("estatus", ""))
                celda = ws.cell(row=fila, column=col, value=valor)
                celda.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[fila].height = 52
            else:
                valor = ev.get(clave, "")
                celda = ws.cell(row=fila, column=col,
                                value=_num(valor) if clave in CAMPOS_MONTO else str(valor))
                if clave in CAMPOS_MONTO:
                    celda.number_format = MONEDA_FMT
    ws.freeze_panes = "A2"


def _fila_kv(ws, fila: int, clave: str, valor, monto: bool = False, negrita: bool = False) -> int:
    c1 = ws.cell(row=fila, column=1, value=clave)
    c2 = ws.cell(row=fila, column=2, value=_num(valor) if monto else valor)
    if monto:
        c2.number_format = MONEDA_FMT
    if negrita:
        c1.font = NEGRITA
        c2.font = NEGRITA
    return fila + 1


def genera_reporte(clasificacion: dict, calculo: dict, ruta_salida: str | Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 22

    isr = calculo["isr"]
    iva = calculo["iva"]
    conc = calculo["conciliacion"]
    t = clasificacion["totales"]

    fila = 1
    titulo = ws.cell(row=fila, column=1,
                     value=f"impuestos AIRBNB mx — Resumen {calculo['periodo']} ({calculo['regimen']}, {calculo['actividad']})")
    titulo.font = Font(bold=True, size=14)
    fila += 2

    # ── Bloque 1: campos del formulario ISR ──────────────────────────────────
    ws.cell(row=fila, column=1,
            value="FORMULARIO ISR — captura en el SAT exactamente estos valores (el skill ya los calculó por ti)").font = NEGRITA
    fila += 1
    fila = _fila_kv(ws, fila, "*Ingresos obtenidos mediante intermediarios",
                    conc["ingresos_recibidos_neto"], monto=True, negrita=True)
    fila = _fila_kv(ws, fila, "*Ingresos obtenidos directamente del usuario  (cobros directos fuera de Airbnb — normalmente 0)",
                    "0.00")
    fila = _fila_kv(ws, fila, "*Retenciones por plataformas tecnológicas",
                    conc["isr_retenido"], monto=True, negrita=True)
    fila = _fila_kv(ws, fila,
                    "  → ISR a cargo que mostrará el SAT (Ingresos × 4%; si Airbnb retuvo exacto = 0)",
                    int(isr["pago_provisional_del_mes_sat"]))
    fila += 1

    # ── Bloque 2: campos del formulario IVA ───────────────────────────────────
    ws.cell(row=fila, column=1,
            value="FORMULARIO IVA — captura en el SAT exactamente estos valores (el skill ya los calculó por ti)").font = NEGRITA
    fila += 1
    fila = _fila_kv(ws, fila, "*Ingresos obtenidos mediante intermediarios",
                    conc["ingresos_recibidos_neto"], monto=True, negrita=True)
    fila = _fila_kv(ws, fila, "*Ingresos obtenidos directamente del usuario  (cobros directos fuera de Airbnb — normalmente 0)",
                    "0.00")
    fila = _fila_kv(ws, fila, "IVA acreditable  ← este valor lo calculó el skill — sin él pagas de más IVA",
                    iva["iva_acreditable_gastos"], monto=True, negrita=True)
    fila = _fila_kv(ws, fila, "*IVA retenido",
                    conc["iva_retenido"], monto=True, negrita=True)
    saldo_ant = iva.get("saldo_favor_iva_anterior", "0")
    fila = _fila_kv(ws, fila, "Acreditamiento saldo a favor periodos anteriores  (solo si el SAT habilita el campo)",
                    saldo_ant, monto=True)
    fila = _fila_kv(ws, fila, "  → Impuesto a cargo IVA estimado (verifica contra el auto-calculado del SAT)",
                    int(iva["a_cargo_sat"]))
    fila = _fila_kv(ws, fila, "  → IVA a favor estimado del mes", iva["a_favor"], monto=True)
    fila += 1

    # ── Nota clave: sin prellenado + deducciones solo en IVA y anual ─────────
    nota = ws.cell(row=fila, column=1,
                   value="⚠ El SAT NO viene prellenado — captura en el portal los valores de este reporte. "
                         "El skill leyó tu CSV de Airbnb y tus facturas del SAT para darte estos números exactos. "
                         "Las deducciones de gastos van al 'IVA acreditable' del formulario IVA (abajo) "
                         "y a tu declaración ANUAL (hojas Deducciones personales e Inversiones).")
    nota.alignment = Alignment(wrap_text=True)
    nota.font = Font(italic=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    ws.row_dimensions[fila].height = 52
    fila += 2

    # ── Bloque 4: detalle ISR ─────────────────────────────────────────────────
    ws.cell(row=fila, column=1, value="DETALLE ISR (acumulado Art. 106 LISR)").font = NEGRITA
    fila += 1
    for clave, etiqueta in [
        ("meses_acumulados", "Meses acumulados"),
        ("ingresos_acumulados", "Ingresos acumulados (base sin IVA)"),
        ("deducciones_acumuladas", "Deducciones acumuladas"),
        ("base_gravable", "Base gravable"),
        ("isr_causado_acumulado", "ISR causado acumulado"),
        ("retenciones_acumuladas", "Retenciones de la plataforma acumuladas"),
        ("pagos_provisionales_previos", "Pagos provisionales previos"),
    ]:
        fila = _fila_kv(ws, fila, etiqueta, isr[clave], monto=clave != "meses_acumulados")
    fila += 1

    # ── Bloque 5: detalle IVA ─────────────────────────────────────────────────
    ws.cell(row=fila, column=1, value="DETALLE IVA (LIVA 1o./4o./5o./6o./18-J)").font = NEGRITA
    fila += 1
    for clave, etiqueta in [
        ("base_gravada", "Base gravada del mes"),
        ("iva_trasladado", "IVA trasladado (causado)"),
        ("iva_retenido_plataforma", "IVA retenido por la plataforma"),
        ("iva_acreditable_gastos", "IVA acreditable de tus gastos"),
        ("saldo_favor_iva_anterior", "Saldo a favor IVA mes anterior"),
    ]:
        fila = _fila_kv(ws, fila, etiqueta, iva.get(clave, "0"), monto=True)
    fila += 1

    # ── Bloque 6: conciliación ────────────────────────────────────────────────
    ws.cell(row=fila, column=1, value="CONCILIACIÓN PLATAFORMA").font = NEGRITA
    fila += 1
    fila = _fila_kv(ws, fila, "ISR retenido según reporte", conc["isr_retenido"], monto=True)
    fila = _fila_kv(ws, fila, "IVA retenido según reporte", conc["iva_retenido"], monto=True)
    fila = _fila_kv(ws, fila, "Base derivada del ISR retenido", conc.get("base_por_isr") or "n/d",
                    monto=conc.get("base_por_isr") is not None)
    fila = _fila_kv(ws, fila, "Base derivada del IVA retenido", conc.get("base_por_iva") or "n/d",
                    monto=conc.get("base_por_iva") is not None)
    fila = _fila_kv(ws, fila, "Depósitos netos recibidos", conc["ingresos_recibidos_neto"], monto=True)
    fila += 1

    ws.cell(row=fila, column=1, value="FACTURAS DEL PERIODO").font = NEGRITA
    fila += 1
    for clave, etiqueta in [
        ("deducibles", "Deducibles"), ("ajustes", "Notas de crédito (ajustes)"),
        ("no_deducibles", "No deducibles"), ("dudosas", "Dudosas sin resolver"),
        ("diferidas", "Diferidas (PPD sin pago en el mes)"),
        ("deducciones_personales", "Candidatas a deducción personal (anual)"),
        ("inversiones", "Inversiones (activo fijo)"),
    ]:
        fila = _fila_kv(ws, fila, etiqueta, t[clave])
    fila += 1

    oportunidad = calculo.get("oportunidad_pagos_definitivos", {})
    if oportunidad:
        ws.cell(row=fila, column=1, value="OPORTUNIDAD DETECTADA").font = NEGRITA
        fila += 1
        fila = _fila_kv(ws, fila, "Ingresos proyectados anuales",
                        oportunidad["ingresos_proyectados_anual"], monto=True)
        nota = ws.cell(row=fila, column=1, value=oportunidad["nota"])
        nota.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        ws.row_dimensions[fila].height = 60
        fila += 2

    for adv in calculo.get("advertencias", []):
        celda = ws.cell(row=fila, column=1, value=f"⚠ {adv}")
        celda.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        ws.row_dimensions[fila].height = 45
        fila += 1
    fila += 1

    disclaimer = ws.cell(row=fila, column=1, value=calculo["disclaimer"])
    disclaimer.font = Font(italic=True)
    disclaimer.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    ws.row_dimensions[fila].height = 60

    evaluaciones = clasificacion["evaluaciones"]

    def _filtra(*estatus: str) -> list[dict]:
        return [e for e in evaluaciones if e["estatus"] in estatus]

    _hoja_facturas(wb, "Deducibles", _filtra("DEDUCIBLE", "AJUSTE"))
    _hoja_facturas(wb, "Dudosas", _filtra("DUDOSA"), add_consejo=True)
    _hoja_facturas(wb, "Rechazadas", _filtra("NO_DEDUCIBLE", "DIFERIDA"), add_consejo=True)
    _hoja_facturas(wb, "Deducciones personales", _filtra("DEDUCCION_PERSONAL"))
    _hoja_facturas(wb, "Inversiones", _filtra("INVERSION"))

    ruta_salida = Path(ruta_salida)
    wb.save(ruta_salida)
    return ruta_salida


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Genera el Excel de resumen mensual")
    parser.add_argument("--clasificacion", required=True, help="JSON de clasificador.py")
    parser.add_argument("--calculo", required=True, help="JSON de calculo_impuestos.py")
    parser.add_argument("-o", "--salida", required=True, help="Ruta del .xlsx de salida")
    args = parser.parse_args(argv)

    clasificacion = json.loads(Path(args.clasificacion).read_text(encoding="utf-8"))
    calculo = json.loads(Path(args.calculo).read_text(encoding="utf-8"))
    ruta = genera_reporte(clasificacion, calculo, args.salida)
    print(f"[reporte] Excel generado: {ruta}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
